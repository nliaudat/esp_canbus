import openpyxl
from openpyxl import Workbook
from openpyxl.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet
from typing import List
import yaml

class Datapoint:
    def __init__(self, row: int, name: str, unit_name: str, unit_id: int, function_group: int, 
                function_number: int, datapoint: int, type_name: str, decimal: int, 
                steps: int, min: int, max: int, writable: bool, unit: str, text: dict[int, str]):
        self.row = row
        self.unit_name = unit_name
        self.unit_id = unit_id
        self.function_group = function_group
        self.function_number = function_number
        self.datapoint = datapoint
        self.type_name = type_name
        self.decimal = decimal
        self.steps = steps
        self.min = min
        self.max = max
        self.writable = writable
        self.unit = unit
        self.text = text
        self.name = name

    def get_id(self) -> str:
        return f'{self.unit_name}_{self.function_group}_{self.function_number}_{self.datapoint}'
    
    def __toptronic_base(self) -> dict:
        return {
            'platform': 'toptronic',
            'name': self.name,
            'device_type': self.unit_name,
            'device_addr': '${TT_' + self.unit_name + '_addr}',
            'function_group': self.function_group,
            'function_number': self.function_number,
            'datapoint': self.datapoint,
        }

    def into_sensor(self) -> dict:
        assert self.type_name != 'LIST'
        
        try:
            decimal_value = int(self.decimal)
        except (ValueError, TypeError):
            decimal_value = 0  # Default value or handle it as you see fit

        decimals = {
            'accuracy_decimals': decimal_value,
            'filters': [{
                'multiply': pow(10, -decimal_value)
            }]
        } if decimal_value > 0  else {} 
        
        
        # convert unit to full device class
        if self.unit is not None and self.unit != '0':
            device_units = unit_to_device_class(self.unit, self.name, class_measurement=True)
        else:
            device_units = {}
                
        # Set sensor internal if writable (inputs)
        internal = { 
            'internal': True,
        } if self.writable else {} 

        return {
            **self.__toptronic_base(),
            'id': self.get_id(),
            'type': self.type_name,
            **decimals,
            **device_units,
            **internal,
        }
      
    
    def into_text_sensor(self) -> dict:
        assert self.type_name == 'LIST'

        return {
            **self.__toptronic_base(),
            'id': self.get_id(),
            'options': [option for option in self.text.values()],
            'values': [value for value in self.text]
        }


    def into_number(self) -> dict:
        assert self.writable and self.type_name != 'LIST'
        
        # convert unit to full device class
        if self.unit is not None and self.unit != '0':
            device_units = unit_to_device_class(self.unit, self.name, class_measurement=False)
        else:
            device_units = {}

        return {
            **self.__toptronic_base(),
            'id': self.get_id()+"_set",
            'type': self.type_name,
            'min_value': self.min,
            'max_value': self.max,
            'step': self.steps,
            'decimal': self.decimal,
            **device_units,
        }


    def into_select(self) -> dict:
        assert self.writable and self.type_name == 'LIST'
        res = self.into_text_sensor()
        res['id'] = res['id'] + '_set'
        return res

class Filter:
    def __init__(self, unit_names: List[str] = None, unit_ids: List[int] = None, rows: List[int] = None):
        self.unit_names = unit_names
        self.unit_ids = unit_ids
        self.rows = rows

    def accepts(self, unit_name: str, unit_id: str, row: int) -> bool:
        if self.unit_names and unit_name not in self.unit_names:
            return False
        if self.unit_ids and unit_id not in self.unit_ids:
            return False
        if self.rows and row not in self.rows:
            return False
        return True

def parse_text(row: List[Cell]) -> dict[int, str]:
    res = {}
    for i, text in enumerate(row[18:]):
        if text.value:
            res[i] = text.value
    return res

def translate(wb: Workbook, datapoints: List[Datapoint], locale: str = 'en') -> None:
    worksheets = {
        'de': wb.worksheets[1],
        'en': wb.worksheets[2],
        'fr': wb.worksheets[3],
        'it': wb.worksheets[4],
    }
    ws = worksheets[locale] or wb.worksheets[2]

    dpMap = {dp.row: dp for dp in datapoints}

    for row in ws.iter_rows(min_row=2):
        if not row[1].value:
            continue

        i = row[1].row
        dp = dpMap.get(i)
        if not dp:
            continue

        dp.name = row[6].value

        if dp.type_name == 'LIST':
            dp.text = parse_text(row)

def parse_datapoints(wb: Workbook, filter: Filter) -> List[Datapoint]:
    ws = wb.worksheets[1]
    datapoints: List[Datapoint] = []
    
    for row in ws.iter_rows(min_row=2):
        unit_name=row[1].value
        if not unit_name:
            continue

        i = row[1].row
        unit_id=row[2].value

        if not filter.accepts(unit_name, unit_id, i):
            continue

        
        datapoint = Datapoint(
            row=i,
            name=row[6].value,
            unit_name=unit_name,
            unit_id=unit_id,
            function_group=row[3].value,
            function_number=row[4].value,
            datapoint=row[5].value,
            type_name=row[8].value,
            decimal=row[9].value,
            steps=row[12].value,
            min=row[13].value,
            max=row[14].value,
            writable=row[15].value == 'Yes',
            unit=row[16].value,
            text={},
        )
        datapoints.append(datapoint)

    return datapoints

def dump_sensors(datapoints: List[Datapoint], path: str):
    sensors = [dp.into_sensor() for dp in datapoints if dp.type_name != 'LIST']
    text_sensors = [dp.into_text_sensor() for dp in datapoints if dp.type_name == 'LIST']
    
    if len(sensors) == 0 and len(text_sensors) == 0:
        return

    all_sensors = {
        **({'sensor': sensors} if len(sensors) > 0 else {}),
        **({'text_sensor': text_sensors} if len(text_sensors) > 0 else {}),
    }

    with open(path, 'w', encoding='utf-8') as f:
        yaml.dump(all_sensors, f, encoding="utf-8", sort_keys=False)

def dump_inputs(datapoints: List[Datapoint], path: str):
    numbers = [dp.into_number() for dp in datapoints if dp.writable and dp.type_name != 'LIST']
    selects = [dp.into_select() for dp in datapoints if dp.writable and dp.type_name == 'LIST']

    if len(numbers) == 0 and len(selects) == 0:
        return

    all_inputs = {
        **({'number': numbers} if len(numbers) > 0 else {}),
        **({'select': selects} if len(selects) > 0 else {}),
    }

    with open(path, 'w', encoding='utf-8') as f:
        yaml.dump(all_inputs, f, encoding="utf-8", sort_keys=False)
        
def unit_to_device_class(unit, name, class_measurement=0):

    # using lambda functions to check if string contains any element from list
    contains_word = lambda s, l: any(map(lambda x: x in s, l))

    state_class = { 
            'state_class': 'measurement',
        } if class_measurement else {} 

    #https://www.home-assistant.io/integrations/number/#device-class
    device_class = {}
    if unit == '%':
        device_class = {
            'unit_of_measurement': '%',
            'icon': 'mdi:percent',
            'device_class': 'power_factor',
            **state_class,
        } 
    if unit == '%' and contains_word(name.lower(), ['humidity', 'humidité', 'feuchte', 'umidit\xE0'] ):
        device_class = {
            'unit_of_measurement': '%',
            'icon': 'mdi:water-percent',
            'device_class': 'humidity',
            **state_class,
        } 
    elif unit == '\xB0C' or unit == '°C':
        device_class = {
            'unit_of_measurement': '°C',
            'icon': 'mdi:temperature-celsius',
            'device_class': 'temperature',
            **state_class,
        } 
    elif unit == 'K' :
        device_class = {
            'unit_of_measurement': '°C',
            'icon': 'mdi:temperature-kelvin',
            'device_class': 'temperature',
            **state_class,
        }     
    elif unit in ['Wh', 'kWh', 'MWh', 'MJ', 'GJ']: 
        device_class = {
            'unit_of_measurement': unit ,
            'icon': 'mdi:flash',
            'device_class': 'energy',
            **state_class,
        }
    elif unit in ['W', 'kW']: 
        device_class = {
            'unit_of_measurement': unit ,
            'icon': 'mdi:lightning-bolt',
            'device_class': 'power',
            **state_class,
        }
    elif unit in ['m³', 'm3', 'L', 'l']: 
        device_class = {
            'unit_of_measurement': unit ,
            'icon': 'mdi:water-percent',
            'device_class': 'volume',
            **state_class,
        }
    elif unit in ['m³/h', 'm3/h', 'L/min', 'l/min']: 
        device_class = {
            'unit_of_measurement': unit ,
            'icon': 'mdi:water-outline',
            'device_class': 'volume_flow_rate',
            **state_class,
        }
    elif unit in ['m³/h', 'm3/h', 'L/min', 'l/min']: 
        device_class = {
            'unit_of_measurement': unit ,
            'icon': 'mdi:water-outline',
            'device_class': 'pressure',
            **state_class,
        }
    return device_class    

if __name__ == "__main__":
    path = 'TTE-GW-Modbus-datapoints.xlsx'
    # filter = Filter(unit_names=['WEZ'], unit_ids=[1])
    # filter = Filter(unit_names=['HV'], unit_ids=[520])
    # filter = Filter(rows=[1382, 3506, 1379])

    # wb = openpyxl.load_workbook(filename=path, read_only=True)

    # datapoints = parse_datapoints(wb, filter)
    # translate(wb, datapoints, 'en')

    # dump_sensors(datapoints, 'sensors.yaml')
    # dump_inputs(datapoints, 'inputs.yaml')
    